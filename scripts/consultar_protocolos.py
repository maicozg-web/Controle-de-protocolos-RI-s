"""
Consulta automática de protocolos no site do Registro de Imóveis de Indaial.

Fluxo:
1. Lê a lista de protocolos da aba "Cadastro" da planilha do Google Sheets
   (com fallback para protocolos.json se o Sheets não estiver configurado).
2. Para cada protocolo, abre o site oficial (riindaial.com.br) com um navegador
   headless (Playwright), preenche código/senha e extrai o resultado.
3. Baixa os PDFs de exigência (quando existirem).
4. Gera uma planilha .xlsx local com o relatório, com formatação e links.
5. Atualiza a aba "Status" do Google Sheets com os mesmos dados, formatada
   (cores por situação, cabeçalho estilizado, links clicáveis de PDF).
6. Envia um e-mail com o relatório e os PDFs anexados.

Variáveis de ambiente esperadas (configuradas como GitHub Secrets):
- GOOGLE_SERVICE_ACCOUNT_JSON : conteúdo do JSON da conta de serviço do Google
- GOOGLE_SHEET_ID             : ID da planilha do Google Sheets
- EMAIL_HOST, EMAIL_PORT      : servidor SMTP (ex: smtp.gmail.com, 587)
- EMAIL_USER, EMAIL_PASS      : credenciais do e-mail remetente (senha de app)
- EMAIL_TO                    : destinatário(s), separados por vírgula
"""

import json
import os
import re
import smtplib
import ssl
from datetime import datetime, date
from email.message import EmailMessage
from pathlib import Path

from playwright.sync_api import sync_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
PROTOCOLOS_FILE = BASE_DIR / "protocolos.json"
OUTPUT_DIR = BASE_DIR / "output"
PDFS_DIR = OUTPUT_DIR / "pdfs"
REPORT_PATH = OUTPUT_DIR / "relatorio_protocolos_ri.xlsx"

SITE_URL = "https://riindaial.com.br/acompanhar-solicitacao/"

NAVY = "1A2B45"
RED_BG = "FBEAE8"
AMBER_BG = "FAF1DE"
GREEN_BG = "EAF2EC"
WHITE = "FFFFFF"

NAVY_RGB = {"red": 0.102, "green": 0.169, "blue": 0.271}
RED_BG_RGB = {"red": 0.984, "green": 0.918, "blue": 0.910}
AMBER_BG_RGB = {"red": 0.980, "green": 0.945, "blue": 0.871}
GREEN_BG_RGB = {"red": 0.918, "green": 0.949, "blue": 0.925}


def col_letter(n):
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def parse_br_date(s):
    if not s:
        return None
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", s)
    if not m:
        return None
    d, mth, y = m.groups()
    return date(int(y), int(mth), int(d))


def _google_client():
    creds_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    sheet_id = os.environ.get("GOOGLE_SHEET_ID")
    if not creds_json or not sheet_id:
        return None, None
    import gspread
    from google.oauth2.service_account import Credentials

    creds_dict = json.loads(creds_json)
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(sheet_id)
    return gc, sh


def carregar_protocolos():
    """Lê a lista de protocolos da aba 'Cadastro' do Sheets; usa protocolos.json como reserva."""
    try:
        _, sh = _google_client()
        if sh is not None:
            ws = sh.worksheet("Cadastro")
            registros = ws.get_all_records(expected_headers=["Protocolo", "Senha", "Referência", "Interessado"])
            protocolos = []
            for r in registros:
                protocolo = str(r.get("Protocolo", "")).strip()
                senha = str(r.get("Senha", "")).strip()
                referencia = str(r.get("Referência", r.get("Referencia", ""))).strip()
                grupo_interessado = str(r.get("Interessado", "")).strip()
                if protocolo and senha:
                    protocolos.append({
                        "protocolo": protocolo,
                        "senha": senha,
                        "referencia": referencia,
                        "grupo_interessado": grupo_interessado,
                    })
            if protocolos:
                print(f"Carregados {len(protocolos)} protocolos da aba 'Cadastro'.")
                return protocolos
            print("[aviso] Aba 'Cadastro' está vazia. Usando protocolos.json local.")
    except Exception as e:
        print(f"[aviso] Não foi possível ler a aba 'Cadastro' ({e}). Usando protocolos.json local.")

    if PROTOCOLOS_FILE.exists():
        return json.loads(PROTOCOLOS_FILE.read_text(encoding="utf-8"))
    return []


def consultar_um_protocolo(page, protocolo, senha):
    """Preenche o formulário do site e extrai os dados do resultado."""
    page.goto(SITE_URL, wait_until="networkidle")

    try:
        page.click("text=Aceitar", timeout=3000)
    except Exception:
        pass

    page.check("input[value='PROTOCOLO']")
    page.check("input[value='NORMAL']")

    codigo_input = page.locator("#ca-codigo")
    senha_input = page.locator("#ca-senha")
    codigo_input.fill("")
    codigo_input.fill(protocolo)
    senha_input.fill("")
    senha_input.fill(senha)

    page.click("button[type='submit']")
    page.wait_for_timeout(1500)

    resultado_texto = ""
    for _tentativa in range(8):
        try:
            resultado_texto = page.locator(".ca-right").inner_text()
        except Exception:
            resultado_texto = ""
        if len(resultado_texto.strip()) > 30:
            break
        page.wait_for_timeout(1500)

    def campo(rotulo):
        m = re.search(rf"{rotulo}\s*\n?\s*([^\n]+)", resultado_texto, re.IGNORECASE)
        return m.group(1).strip() if m else None

    if "não encontrad" in resultado_texto.lower():
        return {"protocolo": protocolo, "erro": "Protocolo não encontrado. Confira código/senha."}

    if len(resultado_texto.strip()) <= 30:
        return {"protocolo": protocolo, "erro": "Não foi possível carregar os dados do site (tempo esgotado). Tente novamente na próxima consulta."}

    dados = {
        "protocolo": protocolo,
        "situacao": campo("SITUAÇÃO \\(ETAPA ATUAL\\)"),
        "cadastro": campo("CADASTRO"),
        "prazoQualificacao": campo("PRAZO QUALIFICAÇÃO"),
        "vencimentoPrenotacao": campo("VENCIMENTO PRENOTAÇÃO"),
        "interessado": campo("INTERESSADO"),
        "solicitante": campo("SOLICITANTE"),
        "natureza": campo("NATUREZA"),
        "total": campo("TOTAL"),
        "exigencias": [],
        "exigencia_links": [],
    }

    exigencia_links = page.locator("a", has_text="Exigência em")
    count = exigencia_links.count()
    for i in range(count):
        el = exigencia_links.nth(i)
        texto = el.inner_text().strip()
        href = el.get_attribute("href")
        dados["exigencias"].append(texto)
        if href:
            dados["exigencia_links"].append(href)

    return dados


def baixar_pdfs_exigencia(page, protocolo, links):
    pasta = PDFS_DIR / protocolo
    pasta.mkdir(parents=True, exist_ok=True)
    caminhos = []
    for i, link in enumerate(links, start=1):
        try:
            with page.expect_download(timeout=15000) as download_info:
                page.goto(link)
            download = download_info.value
            destino = pasta / f"exigencia_{i}.pdf"
            download.save_as(str(destino))
            caminhos.append(str(destino))
        except Exception as e:
            print(f"  [aviso] Não foi possível baixar PDF {i} do protocolo {protocolo}: {e}")
    return caminhos


def rodar_consultas(protocolos_cfg):
    resultados = []
    with sync_playwright() as p:
        browser = p.chromium.launch()
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        for item in protocolos_cfg:
            protocolo = item["protocolo"]
            senha = item["senha"]
            referencia = item.get("referencia", "")
            grupo_interessado = item.get("grupo_interessado", "")
            print(f"Consultando protocolo {protocolo}...")
            dados = consultar_um_protocolo(page, protocolo, senha)
            dados["referencia"] = referencia
            dados["grupo_interessado"] = grupo_interessado

            if dados.get("exigencia_links"):
                dados["pdfs"] = baixar_pdfs_exigencia(page, protocolo, dados["exigencia_links"])
            else:
                dados["pdfs"] = []

            resultados.append(dados)

        browser.close()

    return resultados


def gerar_planilha(resultados):
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    max_exig = max([len(d.get("exigencias", [])) for d in resultados] + [1])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Protocolos RI"

    ncols = 10 + max_exig + 1
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "Painel de Protocolos — Registro de Imóveis"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Atualizado automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color=NAVY)

    headers = ["Protocolo", "Referência", "Interessado", "Solicitante", "Natureza",
               "Situação Atual", "Prazo Qualificação", "Vencimento Prenotação",
               "Total (R$)", "Qtde Exigências"] + [f"Exigência {i+1}" for i in range(max_exig)] + ["Última Consulta"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 32

    today = date.today()
    row = header_row + 1
    for d in resultados:
        if d.get("erro"):
            ws.cell(row=row, column=1, value=d["protocolo"])
            ws.cell(row=row, column=2, value=d.get("referencia", ""))
            ws.cell(row=row, column=6, value=d["erro"]).font = Font(color="A83C34", bold=True)
            row += 1
            continue

        prazo_dt = parse_br_date(d.get("prazoQualificacao"))
        vencido = prazo_dt is not None and prazo_dt < today
        tem_exigencia = len(d.get("exigencias", [])) > 0
        fill = RED_BG if tem_exigencia else (AMBER_BG if vencido else GREEN_BG)

        base_valores = [
            d["protocolo"], d.get("referencia", ""), d.get("interessado", ""),
            d.get("solicitante", ""), d.get("natureza", ""), d.get("situacao", ""),
            d.get("prazoQualificacao", ""), d.get("vencimentoPrenotacao", ""),
            d.get("total", ""), len(d.get("exigencias", [])),
        ]
        for i, v in enumerate(base_valores, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=fill)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        links = d.get("exigencia_links", [])
        for i in range(max_exig):
            col = 11 + i
            c = ws.cell(row=row, column=col)
            c.fill = PatternFill("solid", fgColor=fill)
            if i < len(links):
                c.value = "Baixar PDF"
                c.hyperlink = links[i]
                c.font = Font(name="Arial", size=9, color="1155CC", underline="single")
            else:
                c.font = Font(name="Arial", size=10)

        ultima_col = 11 + max_exig
        c = ws.cell(row=row, column=ultima_col, value=datetime.now().strftime("%d/%m/%Y %H:%M"))
        c.font = Font(name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor=fill)

        row += 1

    widths = [12, 22, 26, 20, 16, 26, 16, 16, 12, 10] + [12] * max_exig + [16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(REPORT_PATH)
    return REPORT_PATH


def esta_pronto_para_retirada(situacao):
    return bool(situacao) and "pronto para retirada" in situacao.lower()


def criar_evento_retirada(referencia, protocolo):
    creds_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    calendar_id = os.environ.get("GOOGLE_CALENDAR_ID")
    if not creds_json or not calendar_id:
        print("[aviso] Google Calendar não configurado — pulando criação de evento.")
        return
    try:
        from google.oauth2.service_account import Credentials
        from google.auth.transport.requests import AuthorizedSession

        creds_dict = json.loads(creds_json)
        scopes = ["https://www.googleapis.com/auth/calendar.events"]
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        session = AuthorizedSession(creds)

        hoje = date.today().isoformat()
        titulo_ref = referencia.strip() if referencia else protocolo
        evento = {
            "summary": f"Retirar Matrículas ({titulo_ref})",
            "start": {"dateTime": f"{hoje}T08:00:00", "timeZone": "America/Sao_Paulo"},
            "end": {"dateTime": f"{hoje}T18:00:00", "timeZone": "America/Sao_Paulo"},
            "description": f"Protocolo {protocolo} pronto para retirada no RI Indaial.",
        }
        url = f"https://www.googleapis.com/calendar/v3/calendars/{calendar_id}/events"
        resp = session.post(url, json=evento, timeout=15)
        if resp.status_code in (200, 201):
            print(f"Evento de retirada criado na agenda para o protocolo {protocolo}.")
        else:
            print(f"[aviso] Falha ao criar evento na agenda ({resp.status_code}): {resp.text[:200]}")
    except Exception as e:
        print(f"[aviso] Erro ao criar evento na agenda: {e}")


def ler_status_anterior(sh):
    """Lê o estado anterior da aba Status para permitir detectar mudanças depois."""
    try:
        ws = sh.worksheet("Status")
    except Exception:
        return {}
    try:
        valores = ws.get_all_values()
    except Exception:
        return {}
    if not valores:
        return {}
    headers = valores[0]
    try:
        idx_protocolo = headers.index("Protocolo")
        idx_situacao = headers.index("Situação Atual")
        idx_qtd = headers.index("Qtde Exigências")
    except ValueError:
        return {}
    anterior = {}
    for linha in valores[1:]:
        if len(linha) <= max(idx_protocolo, idx_situacao, idx_qtd):
            continue
        protocolo = linha[idx_protocolo].strip()
        if not protocolo:
            continue
        anterior[protocolo] = {
            "situacao": linha[idx_situacao].strip(),
            "qtd_exigencias": linha[idx_qtd].strip(),
        }
    return anterior


def atualizar_google_sheets(resultados):
    _, sh = _google_client()
    if sh is None:
        print("[aviso] Credenciais do Google Sheets não configuradas — pulando essa etapa.")
        return []

    import gspread

    anterior = ler_status_anterior(sh)

    try:
        ws = sh.worksheet("Status")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Status", rows=200, cols=20)

    max_exig = max([len(d.get("exigencias", [])) for d in resultados] + [1])
    headers = ["Protocolo", "Referência", "Interessado", "Solicitante", "Natureza",
               "Situação Atual", "Prazo Qualificação", "Vencimento Prenotação",
               "Total (R$)", "Qtde Exigências"] + [f"Exigência {i+1}" for i in range(max_exig)] + ["Última Consulta"]
    ncols = len(headers)

    ws.clear()
    try:
        ws.resize(rows=max(len(resultados) + 10, 20), cols=max(ncols, 10))
    except Exception:
        pass

    hoje = datetime.now().strftime("%d/%m/%Y %H:%M")
    today = date.today()
    linhas = [headers]
    cores_linhas = []

    for d in resultados:
        if d.get("erro"):
            linha = [d["protocolo"], d.get("referencia", ""), "", "", "", d["erro"], "", "", "", 0]
            linha += [""] * max_exig + [hoje]
            cores_linhas.append(AMBER_BG_RGB)
        else:
            prazo_dt = parse_br_date(d.get("prazoQualificacao"))
            vencido = prazo_dt is not None and prazo_dt < today
            tem_exig = len(d.get("exigencias", [])) > 0
            cor = RED_BG_RGB if tem_exig else (AMBER_BG_RGB if vencido else GREEN_BG_RGB)
            cores_linhas.append(cor)

            links = d.get("exigencia_links", [])
            exig_cols = []
            for i in range(max_exig):
                if i < len(links):
                    exig_cols.append(f'=HYPERLINK("{links[i]}","Baixar PDF")')
                else:
                    exig_cols.append("")

            linha = [
                d["protocolo"], d.get("referencia", ""), d.get("interessado", ""),
                d.get("solicitante", ""), d.get("natureza", ""), d.get("situacao", ""),
                d.get("prazoQualificacao", ""), d.get("vencimentoPrenotacao", ""),
                d.get("total", ""), len(d.get("exigencias", [])),
            ] + exig_cols + [hoje]
        linhas.append(linha)

    ws.update(values=linhas, range_name="A1", value_input_option="USER_ENTERED")

    last_col = col_letter(ncols)

    ws.format(f"A1:{last_col}1", {
        "backgroundColor": NAVY_RGB,
        "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
        "horizontalAlignment": "CENTER",
        "wrapStrategy": "WRAP",
    })

    try:
        ws.freeze(rows=1)
    except Exception:
        pass

    for idx, cor in enumerate(cores_linhas, start=2):
        ws.format(f"A{idx}:{last_col}{idx}", {"backgroundColor": cor})

    widths = [90, 190, 210, 170, 130, 210, 130, 140, 90, 100] + [110] * max_exig + [140]
    requests = []
    for i, w in enumerate(widths[:ncols]):
        requests.append({
            "updateDimensionProperties": {
                "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": i, "endIndex": i + 1},
                "properties": {"pixelSize": w},
                "fields": "pixelSize",
            }
        })
    if requests:
        sh.batch_update({"requests": requests})

    print("Google Sheets (aba 'Status') atualizado com formatação.")

    for d in resultados:
        if d.get("erro"):
            continue
        protocolo_ev = str(d.get("protocolo", ""))
        situacao_nova_ev = d.get("situacao") or ""
        info_antiga_ev = anterior.get(protocolo_ev, {})
        situacao_antiga_ev = info_antiga_ev.get("situacao", "")
        if esta_pronto_para_retirada(situacao_nova_ev) and not esta_pronto_para_retirada(situacao_antiga_ev):
            criar_evento_retirada(d.get("referencia", ""), protocolo_ev)

    mudancas = []
    for d in resultados:
        protocolo = str(d.get("protocolo", ""))
        info_antiga = anterior.get(protocolo)
        if info_antiga is None:
            continue
        if d.get("erro"):
            continue
        situacao_antiga = info_antiga.get("situacao", "")
        situacao_nova = d.get("situacao") or ""
        try:
            qtd_antiga = int(info_antiga.get("qtd_exigencias", "0") or "0")
        except ValueError:
            qtd_antiga = 0
        qtd_nova = len(d.get("exigencias", []))

        mudou_situacao = situacao_antiga and situacao_nova and situacao_antiga != situacao_nova
        mudou_exigencia = qtd_nova > qtd_antiga

        if mudou_situacao or mudou_exigencia:
            mudancas.append({
                "protocolo": protocolo,
                "referencia": d.get("referencia", ""),
                "grupo_interessado": d.get("grupo_interessado", ""),
                "situacao_antiga": situacao_antiga,
                "situacao_nova": situacao_nova,
                "qtd_antiga": qtd_antiga,
                "qtd_nova": qtd_nova,
            })

    return mudancas


def enviar_email(resultados, anexo_planilha):
    host = os.environ.get("EMAIL_HOST")
    port = int(os.environ.get("EMAIL_PORT", "587"))
    user = os.environ.get("EMAIL_USER")
    senha = os.environ.get("EMAIL_PASS")
    destinatarios = os.environ.get("EMAIL_TO", "")
    if not all([host, user, senha, destinatarios]):
        print("[aviso] Credenciais de e-mail não configuradas — pulando essa etapa.")
        return

    total_exigencias = sum(len(d.get("exigencias", [])) for d in resultados if not d.get("erro"))
    vencidos = sum(
        1 for d in resultados
        if not d.get("erro") and parse_br_date(d.get("prazoQualificacao"))
        and parse_br_date(d.get("prazoQualificacao")) < date.today()
    )

    msg = EmailMessage()
    msg["Subject"] = f"Relatório de Protocolos RI — {datetime.now().strftime('%d/%m/%Y')}"
    msg["From"] = user
    msg["To"] = destinatarios
    corpo = (
        f"Relatório automático de acompanhamento de protocolos no RI Indaial.\n\n"
        f"Total de protocolos monitorados: {len(resultados)}\n"
        f"Com exigência pendente: {total_exigencias}\n"
        f"Com prazo de qualificação vencido: {vencidos}\n\n"
        f"Planilha completa em anexo. PDFs de exigência (quando houver) também anexados.\n"
        f"A planilha online (Google Sheets) também foi atualizada, aba 'Status'.\n"
    )
    msg.set_content(corpo)

    with open(anexo_planilha, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=anexo_planilha.name,
        )

    for d in resultados:
        for caminho in d.get("pdfs", []):
            p = Path(caminho)
            if p.exists():
                with open(p, "rb") as f:
                    msg.add_attachment(
                        f.read(), maintype="application", subtype="pdf",
                        filename=f"{d['protocolo']}_{p.name}",
                    )

    context = ssl.create_default_context()
    with smtplib.SMTP(host, port) as server:
        server.starttls(context=context)
        server.login(user, senha)
        server.send_message(msg)
    print("E-mail enviado.")


ROTEAMENTO_EMAILS = {
    "GERALOTE": ["Vendas@geralote.com.br", "maicozg@gmail.com"],
    "DLV": ["dlvconstroi@gmail.com", "maicozg@gmail.com"],
    "GREIDE ENGENHARIA": ["lais@greideengenharia.com.br", "maicozg@gmail.com"],
    "GREIDE CONSTRUTORA": ["compras@greideengenharia.com.br", "maicozg@gmail.com"],
}


def destinatarios_para_grupo(grupo_interessado, destinatarios_padrao):
    chave = (grupo_interessado or "").strip().upper()
    return ROTEAMENTO_EMAILS.get(chave, destinatarios_padrao)


def enviar_alerta_mudancas(mudancas):
    if not mudancas:
        print("Nenhuma mudança de status/exigência detectada desde a última consulta.")
        return

    host = os.environ.get("EMAIL_HOST")
    port = int(os.environ.get("EMAIL_PORT", "587"))
    user = os.environ.get("EMAIL_USER")
    senha = os.environ.get("EMAIL_PASS")
    email_padrao_str = os.environ.get("EMAIL_TO", "")
    if not all([host, user, senha]):
        print("[aviso] Credenciais de e-mail não configuradas — pulando alerta de mudanças.")
        return

    destinatarios_padrao = [e.strip() for e in email_padrao_str.split(",") if e.strip()]

    grupos = {}
    for m in mudancas:
        destinatarios = destinatarios_para_grupo(m.get("grupo_interessado"), destinatarios_padrao)
        if not destinatarios:
            continue
        chave = tuple(sorted(set(destinatarios)))
        grupos.setdefault(chave, []).append(m)

    for destinatarios, lista in grupos.items():
        _enviar_um_alerta(host, port, user, senha, list(destinatarios), lista)


def _enviar_um_alerta(host, port, user, senha, destinatarios, mudancas):
    msg = EmailMessage()
    msg["Subject"] = f"Alerta de mudança — Protocolos RI ({len(mudancas)})"
    msg["From"] = user
    msg["To"] = ", ".join(destinatarios)

    linhas = ["Foram detectadas mudanças nos protocolos abaixo, após a consulta mais recente ao RI Indaial:", ""]
    for m in mudancas:
        linhas.append(f"Protocolo {m['protocolo']} ({m['referencia']}):")
        if m["situacao_antiga"] != m["situacao_nova"]:
            linhas.append(f"  - Situação: \"{m['situacao_antiga']}\" -> \"{m['situacao_nova']}\"")
        if m["qtd_nova"] > m["qtd_antiga"]:
            linhas.append(f"  - Exigências: {m['qtd_antiga']} -> {m['qtd_nova']} (nova exigência registrada)")
        linhas.append("")

    linhas.append("Confira os detalhes completos na planilha (aba 'Status') ou no painel web.")
    msg.set_content("\n".join(linhas))

    context = ssl.create_default_context()
    with smtplib.SMTP(host, port) as server:
        server.starttls(context=context)
        server.login(user, senha)
        server.send_message(msg)
    print(f"Alerta de mudança enviado para {', '.join(destinatarios)} ({len(mudancas)} protocolo(s)).")


if __name__ == "__main__":
    protocolos_cfg = carregar_protocolos()
    if not protocolos_cfg:
        print("Nenhum protocolo cadastrado (aba 'Cadastro' vazia e protocolos.json ausente/vazio).")
    else:
        resultados = rodar_consultas(protocolos_cfg)
        planilha = gerar_planilha(resultados)
        mudancas = atualizar_google_sheets(resultados)
        enviar_email(resultados, planilha)
        enviar_alerta_mudancas(mudancas)
    print("Concluído.")
